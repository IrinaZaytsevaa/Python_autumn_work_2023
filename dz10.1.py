import openpyxl
wb = openpyxl.load_workbook('dz1.xlsx')
ws = wb['Prod']

d = {}
for i in range(1, ws.max_row + 1):
    fio = ws.cell(i, 1).value
    sal = ws.cell(i, 2).value
    d[fio] = d.get(fio, 0) + sal

ws = wb['Total']
for i, fio in enumerate(d.keys()):
    ws.cell(i + 1, 1).value = fio
    ws.cell(i + 1, 2).value = d[fio]
mr = ws.max_row + 1
ws.cell(mr, 1).value = 'ИТОГО:'
ws.cell(mr, 2).value = sum(d.values())

wb.save('dz1.xlsx')
